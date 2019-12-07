# This file calculates the angle per cell from the solar array
# Take vertical (and horizontal (?)) angles
# 0 rad/degrees is in reference to the tail of car
from math import pi, atan
from openpyxl import load_workbook

# Appending a file
filename = 'MSXIV-Strategy-Cell-Angles.xlsx'
wb = load_workbook(filename)
ws = wb.active

ws['A1'] = 'Cell ID'
ws['B1'] = 'Angle (Radians)'
ws['C1'] = 'Angle (Degrees)'
# ws['D1'] = 'Notes'

row = int(input("Row Number (Default = 2): "))  # start at this row - editing

try:
    while True:

        # Cell Locations
        # row = 1 # always starts at row 1 of sheet
        cell_id = 'A' + str(row)
        rad_cell_id = 'B' + str(row)
        deg_cell_id = 'C' + str(row)
        # notes

        # inputs
        solar_cell_id = input("Solar Cell ID: ")  # STRING ID of cell
        ws[cell_id] = solar_cell_id

        dy = float(input("dy value - opposite (mm): "))  # mm - in x-y plane
        dz = float(input("dz value - adjacent (mm): "))  # mm

        def to_deg(angle_rad):
            deg = float(angle_rad * 180 / pi)
            return deg

        # Average angle of the panel
        def calculate_angle(dy_val, dz_val):  # calculates angle from normal
            norm_angle = atan(dy/dz)

            print("Normalized Angle:", norm_angle, "radians")
            ws[rad_cell_id] = norm_angle

            print("Normalized Angle:", to_deg(norm_angle), "degrees")
            ws[deg_cell_id] = to_deg(norm_angle)

        calculate_angle(dy, dz)

        # Move to next row
        row = row + 1

except KeyboardInterrupt:  # Ctrl-C
    pass

wb.save(filename=filename)
