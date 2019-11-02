from math import cos, sin, pi, acos, tan, asin
from solar import SolarDay
from cell import SolarCell
from numpy import linspace
from openpyxl import Workbook
from openpyxl import load_workbook

from openpyxl import cell

# Read Solar Cell Angles Sheet
filename = 'MSXIV-Strategy-Cell-Angles.xlsx'
wb = load_workbook(filename, data_only=True)
ws = wb.active

# Create new workbook for Energy Outputs
wb1 = Workbook()
ws1 = wb1.active

# To do:
# Create 3D surface as a combination of all cells: i.e. takes the average energy from all cells
# Input values into the module_angle to account for total energy
# In series/parallel configuration

def to_rad(angle):
    rad = angle * 2 * pi / 360
    return rad

def integrate(x,y):
    # Given an X and Y data set, numerically integrate
    integral = 0
    for i in range(len(x) - 1):
        integral += (x[i + 1] - x[i]) * (y[i + 1] + y[i]) / 2
    return integral

class SolarArray:

    def __init__(self, day, latitude, longitude, timezone, cloudiness):
        self.day = day
        self.lat = latitude
        self.long = longitude
        self.cloud = cloudiness
        self.points = []
        self.time = timezone

    # calculate and store total energy values
    def totalEnergy(self):
        total_energy = 0
        row = 1
        for row in range(2,309):
            for column in "E":
                cell = "{}{}".format(column, row)
                module_angle = ws[cell].value
                ws1['A' + str(row)] = module_angle

                d = SolarDay(self.day, self.lat, self.long, self.time, self.cloud, module_angle)
                # d = SolarDay(day, latitude, longitude, timezone, cloudiness, module_angle)
                insol = integrate(d.energy_received()[1], d.energy_received()[0])
                ws1['B' + str(row)] = insol

                energy = insol * 5 * 0.17
                ws1['C' + str(row)] = energy

                total_energy = total_energy + energy
                row = row + 1
        # print(total_energy)
        wb1.save('energy.xlsx')
        return(total_energy)

if __name__ == '__main__':
    test = SolarArray(182, 30.28, 97.73, 8, 0.5)
    print(test.totalEnergy())
